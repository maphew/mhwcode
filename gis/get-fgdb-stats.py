#!/usr/bin/env uv run --script
#
# /// script
# requires-python = ">=3.13"
# dependencies = [
#     "fiona",
#     "typer",
#     "rich",
#     "openpyxl",
# ]
# ///

"""Scan an FGDB and output a CSV of table metadata of size, geometry type, etc.

Typical usage:
    uv run get-fgdb-stats <gdb_path> [--csv <csv_path>] [--xlsx <xlsx_path>] [--sort-by <column>]
"""
import csv
import datetime
from pathlib import Path
from typing import Optional
import fiona
import xml.etree.ElementTree as ET
import typer
from rich.console import Console
from rich.table import Table
from rich import box
import openpyxl
from openpyxl.utils import get_column_letter

app = typer.Typer(add_completion=False)
console = Console()

def scan_fgdb(gdb_path: Path):
    if not gdb_path.is_dir():
        raise ValueError(f"Not a folder: {gdb_path}")

    console.print(f"[bold blue]Scanning {gdb_path}...[/bold blue]")

    # 1. Read GDB_Items to map PhysicalName -> Logical Name
    physical_to_name = {}
    
    try:
        with fiona.open(str(gdb_path), layer='GDB_Items') as src:
            for feat in src:
                props = feat['properties']
                l_name = props.get('Name')
                definition = props.get('Definition')
                
                if l_name and definition:
                    try:
                        # Parse XML to find DSID
                        root = ET.fromstring(definition)
                        dsid_elem = None
                        for child in root:
                            if child.tag.endswith("DSID"):
                                dsid_elem = child
                                break
                        
                        if dsid_elem is not None:
                            dsid = int(dsid_elem.text)
                            p_name = f"a{dsid:08x}"
                            physical_to_name[p_name] = l_name
                    except Exception:
                        pass
    except Exception as e:
        console.print(f"[bold red]Warning:[/bold red] Could not read GDB_Items directly: {e}")
    
    # 2. Get list of layers and their geometry types using fiona
    layers_info = {}
    try:
        layer_names = fiona.listlayers(str(gdb_path))
        for name in layer_names:
            try:
                with fiona.open(str(gdb_path), layer=name) as src:
                    layers_info[name] = {
                        "geometry": src.schema['geometry'],
                        "type": "File Geodatabase Feature Class"
                    }
            except Exception as e:
                console.print(f"[yellow]Could not open layer {name}: {e}[/yellow]")
                layers_info[name] = {"geometry": "Unknown", "type": "Unknown"}
    except Exception as e:
        console.print(f"[bold red]Error listing layers: {e}[/bold red]")

    # 3. Calculate sizes by grouping physical files
    physical_sizes = {}
    physical_mod_times = {}
    
    for p in gdb_path.iterdir():
        if not p.is_file():
            continue
            
        stem = p.stem.lower()
        if stem not in physical_sizes:
            physical_sizes[stem] = 0
            physical_mod_times[stem] = 0
            
        size = p.stat().st_size
        mtime = p.stat().st_mtime
        
        physical_sizes[stem] += size
        physical_mod_times[stem] = max(physical_mod_times[stem], mtime)

    # 4. Build the final rows
    rows = []
    name_to_physical = {v: k for k, v in physical_to_name.items()}
    all_names = set(physical_to_name.values()) | set(layers_info.keys())
    
    for name in sorted(all_names):
        phys_name = name_to_physical.get(name)
        if not phys_name:
            continue

        phys_key = phys_name.lower()
        size_bytes = physical_sizes.get(phys_key, 0)
        mtime_ts = physical_mod_times.get(phys_key, 0)
        
        if mtime_ts > 0:
            # Match format: 2025-11-24 4:26:41 PM
            # %I is 01-12, %-I is 1-12 (platform dependent, usually # on Windows, - on Unix)
            # Let's stick to standard %I for now or try to match exactly if needed.
            # Expected: 4:26:41 PM. My output: 10:31:41 AM.
            # It seems standard %I:%M:%S %p is fine.
            dt_mod = datetime.datetime.fromtimestamp(mtime_ts).strftime('%Y-%m-%d %I:%M:%S %p')
        else:
            dt_mod = ""
            
        geom = "Unknown"
        ftype = "File Geodatabase Table"
        
        if name in layers_info:
            raw_geom = layers_info[name]['geometry']
            if raw_geom:
                ftype = "File Geodatabase Feature Class"
                # Simplify Geometry
                u_geom = raw_geom.upper()
                if "POINT" in u_geom:
                    geom = "Point"
                elif "LINE" in u_geom:
                    geom = "Line"
                elif "POLYGON" in u_geom:
                    geom = "Polygon"
                else:
                    geom = raw_geom # Fallback
            else:
                 geom = ""

        size_kb = int(round(size_bytes / 1024))
        
        rows.append({
            "Name": name,
            "Type": ftype,
            "Geometry": geom,
            "Date Modified": dt_mod,
            "Size KB": size_kb,
            "Path": str(gdb_path).replace("\\", "/") # Normalize path separators
        })

    return rows

@app.command()
def main(
    gdb_path: Path = typer.Argument(..., help="Path to the File Geodatabase (.gdb)"),
    output_csv: Optional[Path] = typer.Option(None, "--csv", help="Output CSV file path"),
    output_xlsx: Optional[Path] = typer.Option(None, "--xlsx", help="Output Excel file path"),
    sort_by: str = typer.Option("Size", help="Column to sort by (Name, Size, Date Modified)"),
):
    """
    Scan a File Geodatabase and report on feature classes and tables.
    """
    try:
        rows = scan_fgdb(gdb_path)
        
        # Sort
        if sort_by.lower() == "size":
            rows.sort(key=lambda x: x["Size KB"], reverse=True)
        elif sort_by.lower() == "name":
            rows.sort(key=lambda x: x["Name"])
        elif sort_by.lower() == "date modified":
            rows.sort(key=lambda x: x["Date Modified"], reverse=True)

        # Display Table
        table = Table(box=box.SIMPLE)
        table.add_column("Name", style="cyan")
        table.add_column("Type", style="magenta")
        table.add_column("Geometry", style="green")
        table.add_column("Date Modified")
        table.add_column("Size KB", justify="right")
        
        for row in rows:
            table.add_row(
                row["Name"],
                row["Type"],
                str(row["Geometry"]),
                row["Date Modified"],
                f"{row['Size KB']:,}" # Format with commas for display
            )
            
        console.print(table)
        console.print(f"\n[bold]Total items:[/bold] {len(rows)}")

        # Exports
        if output_csv:
            if rows:
                with open(output_csv, "w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(f, fieldnames=rows[0].keys())
                    w.writeheader()
                    w.writerows(rows)
                console.print(f"[green]Saved CSV to {output_csv}[/green]")

        if output_xlsx:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GDB Inventory"
            
            # Headers
            headers = ["Name", "Type", "Geometry", "Date Modified", "Size KB", "Path"]
            ws.append(headers)
            
            # Data
            for row in rows:
                ws.append([
                    row["Name"],
                    row["Type"],
                    row["Geometry"],
                    row["Date Modified"],
                    row["Size KB"],
                    row["Path"]
                ])
                
            # Formatting
            # Make header bold
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                
            # Auto-filter
            ws.auto_filter.ref = ws.dimensions
            
            # Column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(output_xlsx)
            console.print(f"[green]Saved Excel to {output_xlsx}[/green]")

    except Exception as e:
        console.print(f"[bold red]Failed:[/bold red] {e}")
        raise typer.Exit(code=1)

if __name__ == "__main__":
    app()
